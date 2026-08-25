"""Purpose: Provides the benchmark excel accuracy command-line utility.

Used by: Run manually or via python -m backend.scripts.benchmark_excel_accuracy.
"""

from __future__ import annotations

import argparse
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scripts.ingest_verified_incidents import DEFAULT_SOURCE, load_records


FIELDS = (
    "domain", "subdomain", "safety_impact", "business_continuity",
    "damage_to_assets", "reputational_impact", "vip_safety_impact",
    "likelihood_of_more_severe_outcomes", "hipo_classification",
)


def score(value: Any) -> int | None:
    if value is None:
        return None
    try:
        number = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None
    return number if number in range(1, 6) else None


def hipo(value: Any) -> str | None:
    if isinstance(value, dict):
        value = value.get("hipo_classification") or value.get("classification")
    label = str(value or "").strip().lower().replace("_", "-")
    if label == "hipo":
        return "HIPO"
    if label in {"non-hipo", "not hipo", "non hipo"}:
        return "Non-HIPO"
    return None


def actual(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "domain": record.get("domain"),
        "subdomain": record.get("subdomain"),
        "safety_impact": score(record.get("safety_impact")),
        "business_continuity": score(record.get("business_continuity")),
        "damage_to_assets": score(record.get("damage_to_assets")),
        "reputational_impact": score(record.get("reputational_impact")),
        "vip_safety_impact": score(record.get("vip_safety")),
        "likelihood_of_more_severe_outcomes": score(record.get("likelihood_of_more_severe_outcome")),
        "hipo_classification": hipo(record.get("hipo_classification")),
    }


def predicted(response: dict[str, Any]) -> dict[str, Any]:
    result = response["result"]
    return {
        "domain": result.get("domain"),
        "subdomain": result.get("subdomain"),
        "safety_impact": score(result.get("safety_impact")),
        "business_continuity": score(result.get("business_continuity")),
        "damage_to_assets": score(result.get("damage_to_assets")),
        "reputational_impact": score(result.get("reputational_impact")),
        "vip_safety_impact": score(result.get("vip_safety_impact")),
        "likelihood_of_more_severe_outcomes": score(result.get("likelihood_of_more_severe_outcomes")),
        "hipo_classification": hipo(result.get("hipo_classification")),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare model answers directly with an Excel answer key.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=Path("outputs/model_accuracy_report.xlsx"))
    parser.add_argument("--api-url", default="http://localhost:8000")
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    records, errors = load_records(args.source)
    if errors:
        raise SystemExit("\n".join(errors))
    if args.limit:
        records = records[: args.limit]

    results = []
    with httpx.Client(timeout=180) as client:
        for index, record in enumerate(records, 1):
            started = time.perf_counter()
            try:
                response = client.post(
                    f"{args.api_url.rstrip('/')}/incident/workflow/start",
                    json={"incident_text": record["incident_summary"]},
                )
                response.raise_for_status()
                prediction = predicted(response.json())
                error = None
            except Exception as exc:
                prediction = {field: None for field in FIELDS}
                error = str(exc)
            answer = actual(record)
            comparisons = {
                field: prediction[field] == answer[field]
                if prediction[field] is not None and answer[field] is not None else None
                for field in FIELDS
            }
            results.append({
                "case_id": record.get("incident_no"),
                "incident_narrative": record["incident_summary"],
                "actual": answer,
                "predicted": prediction,
                "match": comparisons,
                "latency_seconds": round(time.perf_counter() - started, 2),
                "error": error,
            })
            print(f"Completed {index}/{len(records)}")

    workbook = Workbook()
    details = workbook.active
    details.title = "Case Comparison"
    headers = ["Case ID", "Incident Narrative"]
    for field in FIELDS:
        label = field.replace("_", " ").title()
        headers.extend([f"Actual {label}", f"Predicted {label}", f"{label} Match"])
    headers.extend(["All Fields Match", "Latency Seconds", "Error"])
    details.append(headers)
    for item in results:
        row = [item["case_id"], item["incident_narrative"]]
        valid_matches = []
        for field in FIELDS:
            match = item["match"][field]
            row.extend([item["actual"][field], item["predicted"][field], match])
            if match is not None:
                valid_matches.append(match)
        row.extend([all(valid_matches) if valid_matches else None, item["latency_seconds"], item["error"]])
        details.append(row)

    summary = workbook.create_sheet("Accuracy Summary")
    summary.append(["Metric", "Correct", "Compared", "Accuracy"])
    for field in FIELDS:
        matches = [item["match"][field] for item in results if item["match"][field] is not None]
        correct = sum(matches)
        summary.append([field.replace("_", " ").title(), correct, len(matches), correct / len(matches) if matches else None])
    pair_matches = [
        item["match"]["domain"] and item["match"]["subdomain"]
        for item in results
        if item["match"]["domain"] is not None and item["match"]["subdomain"] is not None
    ]
    summary.append(["Domain + Subdomain Pair", sum(pair_matches), len(pair_matches), sum(pair_matches) / len(pair_matches) if pair_matches else None])
    all_matches = [all(value for value in item["match"].values() if value is not None) for item in results if not item["error"]]
    summary.append(["All Fields Exact", sum(all_matches), len(all_matches), sum(all_matches) / len(all_matches) if all_matches else None])

    fill = PatternFill("solid", fgColor="0B6E4F")
    for sheet in (details, summary):
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    details.column_dimensions["B"].width = 60
    for index in range(1, len(headers) + 1):
        if index != 2:
            details.column_dimensions[get_column_letter(index)].width = 18
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["D"].width = 16
    for cell in summary["D"][1:]:
        cell.number_format = "0.0%"

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Saved accuracy report: {args.output.resolve()}")


if __name__ == "__main__":
    main()
