"""Purpose: Provides the benchmark pdf accuracy command-line utility.

Used by: Run manually or via python -m backend.scripts.benchmark_pdf_accuracy.
"""

from __future__ import annotations

import argparse
import json
import re
import time
from pathlib import Path
from typing import Any

import fitz
import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scripts.benchmark_excel_accuracy import FIELDS, actual, predicted


DEFAULT_SOURCE = Path(
    r"C:\Users\tvish\Downloads\HIPO_100_Labelled_Accuracy_Test_Cases.pdf"
)
DEFAULT_OUTPUT = Path("outputs/hipo_unseen_accuracy_100.xlsx")

CASE_PATTERN = re.compile(
    r"Expected Domain\n(.+?)\nCode\n(.+?)\nExpected Sub-domain\n(.+?)\n"
    r"Expected HIPO\n(YES|NO)\nDetailed incident narrative\n(.*?)\n"
    r"Safety\nBC\nAssets\nReputation\nVIP\nLikelihood\nTotal\n"
    r"(\d+)\n(\d+)\n(\d+)\n(\d+)\n(\d+)\n(\d+)\n(\d+)",
    re.DOTALL,
)


def parse_case_text(text: str, page_number: int) -> dict[str, Any]:
    case_match = re.search(r"\bNEW\d{3}\b", text)
    match = CASE_PATTERN.search(text.replace("\r", ""))
    if case_match is None or match is None:
        raise ValueError(f"Page {page_number}: case layout could not be parsed")

    domain, code, subdomain, expected_hipo, narrative, *raw_numbers = match.groups()
    numbers = [int(value) for value in raw_numbers]
    safety, continuity, assets, reputation, vip, likelihood, total = numbers
    scores = (safety, continuity, assets, reputation, vip, likelihood)
    if any(value not in range(1, 6) for value in scores):
        raise ValueError(f"Page {page_number}: scores must be integers from 1 to 5")
    if sum(scores) != total:
        raise ValueError(
            f"Page {page_number}: total {total} does not equal score sum {sum(scores)}"
        )

    hipo_label = "HIPO" if expected_hipo == "YES" else "Non-HIPO"
    rule_hipo = max(scores[:5]) >= 4 and likelihood >= 4
    if rule_hipo != (hipo_label == "HIPO"):
        raise ValueError(f"Page {page_number}: label conflicts with the coupled HIPO rule")

    return {
        "incident_no": case_match.group(),
        "incident_summary": " ".join(narrative.split()),
        "domain": domain.strip(),
        "subdomain": subdomain.strip(),
        "code": code.strip(),
        "safety_impact": safety,
        "business_continuity": continuity,
        "damage_to_assets": assets,
        "reputational_impact": reputation,
        "vip_safety": vip,
        "likelihood_of_more_severe_outcome": likelihood,
        "total_score": total,
        "hipo_classification": hipo_label,
        "source_page": page_number,
    }


def load_pdf_records(source: Path) -> list[dict[str, Any]]:
    if not source.exists():
        raise FileNotFoundError(f"PDF test set not found: {source}")
    document = fitz.open(source)
    try:
        records = [
            parse_case_text(document[index].get_text("text"), index + 1)
            for index in range(1, document.page_count)
        ]
    finally:
        document.close()
    case_ids = [record["incident_no"] for record in records]
    if len(case_ids) != len(set(case_ids)):
        raise ValueError("The PDF contains duplicate case identifiers")
    if not records:
        raise ValueError("The PDF contains no labelled cases")
    return records


def compare(record: dict[str, Any], prediction: dict[str, Any], *, latency: float,
            error: str | None, attempts: int,
            diagnostics: dict[str, Any] | None = None) -> dict[str, Any]:
    answer = actual(record)
    matches = {
        field: prediction[field] == answer[field]
        if prediction[field] is not None and answer[field] is not None else None
        for field in FIELDS
    }
    return {
        "case_id": record["incident_no"],
        "incident_narrative": record["incident_summary"],
        "actual": answer,
        "predicted": prediction,
        "match": matches,
        "latency_seconds": round(latency, 2),
        "attempts": attempts,
        "error": error,
        "diagnostics": diagnostics or {},
    }


def write_report(results: list[dict[str, Any]], output: Path) -> None:
    workbook = Workbook()
    details = workbook.active
    details.title = "Case Comparison"
    headers = ["Case ID", "Incident Narrative"]
    for field in FIELDS:
        label = field.replace("_", " ").title()
        headers.extend([f"Actual {label}", f"Predicted {label}", f"{label} Match"])
    headers.extend([
        "All Fields Match", "Latency Seconds", "Attempts", "Error",
        "HIPO Assessment Mode", "HIPO Assessment Provider", "HIPO Facts Provider",
        "HIPO Narrative Corrections", "HIPO Event Profile", "HIPO Event Evidence",
        "HIPO Event Corrections", "HIPO Stage Timings Ms", "HIPO Review Required",
        "HIPO Missing Information",
    ])
    details.append(headers)

    for item in results:
        row = [item["case_id"], item["incident_narrative"]]
        valid_matches = []
        for field in FIELDS:
            match = item["match"][field]
            row.extend([item["actual"][field], item["predicted"][field], match])
            if match is not None:
                valid_matches.append(match)
        row.extend([
            all(valid_matches) if valid_matches else None,
            item["latency_seconds"], item["attempts"], item["error"],
            item.get("diagnostics", {}).get("hipo_assessment_mode"),
            item.get("diagnostics", {}).get("hipo_assessment_provider"),
            item.get("diagnostics", {}).get("hipo_facts_provider"),
            "; ".join(item.get("diagnostics", {}).get("hipo_narrative_corrections") or []),
            item.get("diagnostics", {}).get("hipo_event_profile"),
            "; ".join(item.get("diagnostics", {}).get("hipo_event_evidence") or []),
            "; ".join(item.get("diagnostics", {}).get("hipo_event_corrections") or []),
            json.dumps(item.get("diagnostics", {}).get("hipo_stage_timings_ms") or {}),
            item.get("diagnostics", {}).get("hipo_review_required"),
            "; ".join(item.get("diagnostics", {}).get("hipo_missing_information") or []),
        ])
        details.append(row)

    summary = workbook.create_sheet("Accuracy Summary")
    summary.append(["Metric", "Correct", "Compared", "Accuracy"])
    for field in FIELDS:
        matches = [
            item["match"][field] for item in results
            if item["match"][field] is not None
        ]
        correct = sum(matches)
        summary.append([
            field.replace("_", " ").title(), correct, len(matches),
            correct / len(matches) if matches else None,
        ])
    pair_matches = [
        item["match"]["domain"] and item["match"]["subdomain"]
        for item in results
        if item["match"]["domain"] is not None
        and item["match"]["subdomain"] is not None
    ]
    summary.append([
        "Domain + Subdomain Pair", sum(pair_matches), len(pair_matches),
        sum(pair_matches) / len(pair_matches) if pair_matches else None,
    ])

    hipo_pairs = [
        (item["actual"]["hipo_classification"], item["predicted"]["hipo_classification"])
        for item in results if item["predicted"]["hipo_classification"] is not None
    ]
    tp = sum(actual_label == predicted_label == "HIPO" for actual_label, predicted_label in hipo_pairs)
    fp = sum(actual_label == "Non-HIPO" and predicted_label == "HIPO" for actual_label, predicted_label in hipo_pairs)
    fn = sum(actual_label == "HIPO" and predicted_label == "Non-HIPO" for actual_label, predicted_label in hipo_pairs)
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    summary.append(["HIPO Precision", tp, tp + fp, precision])
    summary.append(["HIPO Recall", tp, tp + fn, recall])
    summary.append(["HIPO F1", None, None, f1])
    summary.append(["Errors", sum(bool(item["error"]) for item in results), len(results), None])

    fill = PatternFill("solid", fgColor="0B6E4F")
    for sheet in (details, summary):
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    details.column_dimensions["B"].width = 70
    for index in range(1, len(headers) + 1):
        if index != 2:
            details.column_dimensions[get_column_letter(index)].width = 18
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["D"].width = 16
    for cell in summary["D"][1:]:
        cell.number_format = "0.0%"

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def save_checkpoint(path: Path, results: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)


def load_checkpoint(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("Checkpoint must contain a JSON list")
    return data


def request_prediction(api_url: str, incident_text: str, max_attempts: int,
                       timeout_seconds: float) -> tuple[dict[str, Any], dict[str, Any], str | None, int]:
    last_error: str | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            response = httpx.post(
                f"{api_url.rstrip('/')}/incident/workflow/start",
                json={"incident_text": incident_text}, timeout=timeout_seconds,
            )
            if response.status_code == 429 or response.status_code >= 500:
                raise httpx.HTTPStatusError(
                    f"HTTP {response.status_code}: {response.text[:1000]}",
                    request=response.request, response=response,
                )
            response.raise_for_status()
            payload = response.json()
            return predicted(payload), payload.get("result", {}).get("diagnostics", {}), None, attempt
        except (httpx.TransportError, httpx.HTTPStatusError) as exc:
            last_error = str(exc)
            retryable = isinstance(exc, httpx.TransportError) or (
                isinstance(exc, httpx.HTTPStatusError)
                and exc.response.status_code in {429, 500, 502, 503, 504}
            )
            if not retryable or attempt == max_attempts:
                break
            retry_after = exc.response.headers.get("Retry-After") if isinstance(
                exc, httpx.HTTPStatusError
            ) else None
            try:
                wait_seconds = float(retry_after) if retry_after else min(
                    5 * (2 ** (attempt - 1)), 60
                )
            except ValueError:
                wait_seconds = min(5 * (2 ** (attempt - 1)), 60)
            time.sleep(max(wait_seconds, 1))
    return {field: None for field in FIELDS}, {}, last_error, max_attempts


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Evaluate the incident pipeline against the labelled HIPO PDF."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--checkpoint", type=Path)
    parser.add_argument("--api-url", default="http://localhost:8000")
    parser.add_argument("--limit", type=int)
    parser.add_argument("--case-id", help="Run one labelled case ID, for example NEW014")
    parser.add_argument(
        "--start-case", type=int, default=1,
        help="One-based case position to start from; 26 starts with the 26th PDF case.",
    )
    parser.add_argument("--max-attempts", type=int, default=2)
    parser.add_argument("--timeout-seconds", type=float, default=240)
    parser.add_argument(
        "--delay-seconds", type=float, default=0,
        help="Pause between cases to avoid hosted API rate limits.",
    )
    args = parser.parse_args()

    records = load_pdf_records(args.source)
    if args.case_id:
        requested_case = args.case_id.strip().upper()
        records = [record for record in records if record["incident_no"] == requested_case]
        if not records:
            raise ValueError(f"Case ID not found in PDF: {requested_case}")
    else:
        if args.start_case < 1:
            raise ValueError("--start-case must be at least 1")
        records = records[args.start_case - 1:]
        if args.limit:
            records = records[:args.limit]
    checkpoint = args.checkpoint or args.output.with_suffix(".checkpoint.json")
    checkpoint_results = load_checkpoint(checkpoint)
    failed_checkpoint_cases = sum(bool(item.get("error")) for item in checkpoint_results)
    results = [item for item in checkpoint_results if not item.get("error")]
    completed = {str(item.get("case_id")) for item in results}
    records = [record for record in records if record["incident_no"] not in completed]
    total = len(results) + len(records)

    print(f"Validated PDF cases: {total}")
    print(f"Resuming after: {len(results)} completed cases")
    if failed_checkpoint_cases:
        print(f"Retrying failed checkpoint cases: {failed_checkpoint_cases}")
    for record in records:
        started = time.perf_counter()
        prediction, diagnostics, error, attempts = request_prediction(
            args.api_url, record["incident_summary"],
            max(1, args.max_attempts), args.timeout_seconds,
        )
        results.append(compare(
            record, prediction, latency=time.perf_counter() - started,
            error=error, attempts=attempts, diagnostics=diagnostics,
        ))
        save_checkpoint(checkpoint, results)
        write_report(results, args.output)
        print(f"Completed {len(results)}/{total}")
        if args.delay_seconds > 0 and len(results) < total:
            time.sleep(args.delay_seconds)

    write_report(results, args.output)
    print(f"Saved accuracy report: {args.output.resolve()}")
    print(f"Saved checkpoint: {checkpoint.resolve()}")


if __name__ == "__main__":
    main()
