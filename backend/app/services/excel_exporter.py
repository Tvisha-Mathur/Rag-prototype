"""Purpose: Implements the excel exporter application service.

Used by: Imported by the incident-analysis runtime, workflow, or supporting evaluation tools.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.app.services.accuracy_evaluator import SCORE_FIELDS, normalize_hipo


EXPORT_FIELDS = (
    "domain", "subdomain", *SCORE_FIELDS, "hipo_classification"
)


def scalar_hipo(value: Any) -> Any:
    try:
        return normalize_hipo(value)
    except ValueError:
        return None


def flatten_result(result: dict[str, Any] | None, prefix: str) -> dict[str, Any]:
    result = result or {}
    flattened: dict[str, Any] = {}
    for field in EXPORT_FIELDS:
        value = result.get(field)
        if isinstance(value, dict):
            value = value.get("score") if field != "hipo_classification" else scalar_hipo(value)
        if field == "hipo_classification":
            value = scalar_hipo(value)
        flattened[f"{prefix}_{field}"] = value
    return flattened


def build_response_workbook(records: list[dict[str, Any]]) -> BytesIO:
    columns = [
        "Factual Summary",
        "Date",
        "Time",
        "Domain",
        "Subdomain",
        "Affected Party Details",
        "Actual / Near Miss",
        "Safety Impact",
        "Business Continuity",
        "Damage to Assets",
        "Reputational Impact",
        "VIP Safety Impact",
        "HIPO Classification",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Incident Responses"
    sheet.append(columns)

    for record in records:
        prediction = record.get("model_prediction") or {}
        row = {
            "Factual Summary": prediction.get("factual_summary"),
            "Date": prediction.get("date"),
            "Time": prediction.get("time"),
            "Domain": prediction.get("domain"),
            "Subdomain": prediction.get("subdomain"),
            "Affected Party Details": prediction.get("affected_party_details"),
            "Actual / Near Miss": prediction.get("actual_or_near_miss"),
            "Safety Impact": prediction.get("safety_impact"),
            "Business Continuity": prediction.get("business_continuity"),
            "Damage to Assets": prediction.get("damage_to_assets"),
            "Reputational Impact": prediction.get("reputational_impact"),
            "VIP Safety Impact": prediction.get("vip_safety_impact"),
            "HIPO Classification": scalar_hipo(prediction.get("hipo_classification")),
        }
        sheet.append([row.get(column) for column in columns])

    header_fill = PatternFill("solid", fgColor="0B6E4F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        width = 18
        if column == "Factual Summary":
            width = 60
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    notes = workbook.create_sheet("Export Information")
    notes.append(["Field", "Value"])
    notes.append(["Generated at (UTC)", datetime.now(UTC).isoformat()])
    notes.append(["Records exported", len(records)])
    notes.append(["Accuracy eligibility", "Requires model_prediction and verified expert_review"])
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 70

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def load_export_records(database: Any, verified_only: bool = False) -> list[dict[str, Any]]:
    match: dict[str, Any] = {"model_prediction": {"$exists": True}}
    if verified_only:
        match.update({"review_status": "verified", "expert_review": {"$exists": True}})
    return list(database["incident_responses"].aggregate([
        {"$match": match},
        {"$lookup": {
            "from": "incident_queries", "localField": "query_id",
            "foreignField": "query_id", "as": "query",
        }},
        {"$unwind": {"path": "$query", "preserveNullAndEmptyArrays": True}},
        {"$set": {"incident_text": "$query.incident_text"}},
        {"$sort": {"created_at": -1}},
    ]))
